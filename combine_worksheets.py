import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
import sys 

def validate_required_files(directory):
    """Validate that all required files are present before processing"""
    try:
        # Check for UUID file
        if not glob.glob(os.path.join(directory, "????????-????-????-????-????????????.xlsx")):
            raise FileNotFoundError("UUID file not found")
            
        # Check for Jobs Report
        if not glob.glob(os.path.join(directory, "Copy of Jobs Report for Performance -DE2_Dated *.xlsx")):
            raise FileNotFoundError("Jobs Report file not found")
            
        # Check for Tech Department file
        if not glob.glob(os.path.join(directory, "Technician Department_Dated *.xlsx")):
            raise FileNotFoundError("Tech Department file not found")
            
        # Check for Time Off file
        time_off_file = os.path.join(directory, "Approved_Time_Off 2023.xlsx")
        if not os.path.exists(time_off_file):
            raise FileNotFoundError("Time Off file not found")
            
        # Note: TGL file is optional, so we don't error if it's missing
        return True
        
    except FileNotFoundError as e:
        print(f"\nError: {str(e)}")
        print("Please ensure all required files are in the directory before running.")
        sys.exit(1)

def find_latest_files(directory):
    try:
        uuid_file = glob.glob(os.path.join(directory, "????????-????-????-????-????????????.xlsx"))[0]
    except IndexError:
        raise FileNotFoundError("UUID file not found")
    
    try:
        jobs_file = max(glob.glob(os.path.join(directory, "Copy of Jobs Report for Performance -DE2_Dated *.xlsx")))
    except ValueError:
        raise FileNotFoundError("Jobs Report file not found")
    
    try:
        tech_file = max(glob.glob(os.path.join(directory, "Technician Department_Dated *.xlsx")))
    except ValueError:
        raise FileNotFoundError("Tech Department file not found")
    
    time_off_file = os.path.join(directory, "Approved_Time_Off 2023.xlsx")
    if not os.path.exists(time_off_file):
        raise FileNotFoundError("Time Off file not found")
    
    # Corrected pattern for TGL file
    tgl_files = glob.glob(os.path.join(directory, "TGLs Set _Dated *.xlsx"))
    tgl_file = max(tgl_files) if tgl_files else None
    
    return {
        'uuid': uuid_file,
        'jobs': jobs_file,
        'tech': tech_file,
        'time_off': time_off_file,
        'tgl': tgl_file
    }

def autofit_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def combine_workbooks(directory, output_file):
    validate_required_files(directory)
    files = find_latest_files(directory)
    
    # Start by copying the UUID file as our base
    shutil.copy2(files['uuid'], output_file)
    
    # Open target workbook
    target_wb = load_workbook(filename=output_file)
    
    # AutoFit UUID file sheets
    for sheet in target_wb.sheetnames:
        autofit_columns(target_wb[sheet])
    
    # Copy from Jobs Report
    source_wb = load_workbook(filename=files['jobs'], data_only=True)
    source_ws = source_wb['Sheet1']
    if 'Sheet1' in target_wb.sheetnames:
        target_wb.remove(target_wb['Sheet1'])
    target_wb.create_sheet('Sheet1')
    target_ws = target_wb['Sheet1']
    
    for row in source_ws:
        for cell in row:
            target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    autofit_columns(target_ws)
    
    # Copy from Tech Department
    source_wb = load_workbook(filename=files['tech'], data_only=True)
    source_ws = source_wb['Sheet1']
    target_wb.create_sheet('Sheet1_Tech')
    target_ws = target_wb['Sheet1_Tech']
    
    for row in source_ws:
        for cell in row:
            target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    autofit_columns(target_ws)
    
    # Copy from Time Off
    source_wb = load_workbook(filename=files['time_off'], data_only=True)
    source_ws = source_wb['2024']
    if '2024' in target_wb.sheetnames:
        target_wb.remove(target_wb['2024'])
    target_wb.create_sheet('2024')
    target_ws = target_wb['2024']
    
    for row in source_ws:
        for cell in row:
            target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
    autofit_columns(target_ws)
    
    # Copy from TGLs Set if file exists
    if files['tgl']:
        source_wb = load_workbook(filename=files['tgl'], data_only=True)
        source_ws = source_wb['Sheet1']
        if 'Sheet1_TGL' in target_wb.sheetnames:
            target_wb.remove(target_wb['Sheet1_TGL'])
        target_wb.create_sheet('Sheet1_TGL')
        target_ws = target_wb['Sheet1_TGL']
        
        for row in source_ws:
            for cell in row:
                target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        autofit_columns(target_ws)
    
    # Save the workbook
    target_wb.save(output_file)

# Usage
directory = r'C:\Users\abatlouni\Downloads'
output_file = os.path.join(directory, 'combined_data.xlsx')

try:
    combine_workbooks(directory, output_file)
    print("Workbook combination completed successfully!")
    if not glob.glob(os.path.join(directory, "TGLs Set _Dated *.xlsx")):
        print("Note: TGLs Set file was not found - skipped this sheet")
except FileNotFoundError as e:
    print(f"Error: {e}")
except Exception as e:
    print(f"An unexpected error occurred: {e}")